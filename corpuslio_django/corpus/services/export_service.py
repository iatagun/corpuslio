"""
Export Service for Corpus Platform
Handles export of concordance, frequency, and ngram data with watermarking.
"""

import csv
import json
import io
from datetime import datetime
from typing import List, Dict, Any, Optional
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment


class ExportService:
    """
    Service class for exporting corpus data with watermarks and citations.
    Supports CSV, JSON, and Excel formats.
    """
    
    def __init__(self, user, document=None, query_text: str = ""):
        """
        Initialize export service.
        
        Args:
            user: Django User object
            document: Document object (optional)
            query_text: Search query that generated results
        """
        self.user = user
        self.document = document
        self.query_text = query_text
        self.timestamp = timezone.now()
    
    def generate_citation(self) -> str:
        """
        Generate citation text for watermark.
        
        Returns:
            Citation string
        """
        # Get document title from metadata or filename
        if self.document:
            doc_title = self.document.filename
            if hasattr(self.document, 'metadata') and isinstance(self.document.metadata, dict):
                doc_title = self.document.metadata.get('title', self.document.filename)
        else:
            doc_title = "全 Corpus"
        
        citation = (
            f"CorpusLIO National Corpus Platform\n"
            f"Exported by: {self.user.username}\n"
            f"Date: {self.timestamp.strftime('%Y-%m-%d %H:%M:%S')}\n"
            f"Document: {doc_title}\n"
        )
        
        if self.query_text:
            citation += f"Query: {self.query_text}\n"
        
        citation += (
            f"\n"
            f"Citation: CorpusLIO Corpus ({self.timestamp.year}). "
            f"Retrieved from CorpusLIO Platform on {self.timestamp.strftime('%B %d, %Y')}.\n"
            f"Please cite this source in academic publications."
        )
        
        return citation
    
    def export_concordance_csv(self, results: List[Dict[str, Any]]) -> bytes:
        """
        Export concordance results as CSV with watermark.
        
        Args:
            results: List of concordance results
                     Each dict should have: left_context, keyword, right_context, document, position
        
        Returns:
            CSV file content as bytes
        """
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Watermark header
        citation = self.generate_citation()
        for line in citation.split('\n'):
            writer.writerow([f"# {line}"])
        writer.writerow([])  # Empty row separator
        
        # Column headers
        writer.writerow(['Left Context', 'Keyword', 'Right Context', 'Document', 'Position'])
        
        # Data rows
        for result in results:
            writer.writerow([
                result.get('left_context', ''),
                result.get('keyword', ''),
                result.get('right_context', ''),
                result.get('document', ''),
                result.get('position', '')
            ])
        
        # Footer watermark
        writer.writerow([])
        writer.writerow([f"Total results: {len(results)}"])
        writer.writerow([f"Export ID: {self.timestamp.strftime('%Y%m%d%H%M%S')}"])
        
        return output.getvalue().encode('utf-8-sig')  # BOM for Excel compatibility
    
    def export_frequency_csv(self, results: List[Dict[str, Any]]) -> bytes:
        """
        Export frequency table as CSV with watermark.
        
        Args:
            results: List of frequency data
                     Each dict should have: word, lemma, pos, frequency, percentage
        
        Returns:
            CSV file content as bytes
        """
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Watermark header
        citation = self.generate_citation()
        for line in citation.split('\n'):
            writer.writerow([f"# {line}"])
        writer.writerow([])
        
        # Column headers
        writer.writerow(['Word', 'Lemma', 'POS', 'Frequency', 'Percentage'])
        
        # Data rows
        for result in results:
            writer.writerow([
                result.get('word', ''),
                result.get('lemma', ''),
                result.get('pos', ''),
                result.get('frequency', 0),
                f"{result.get('percentage', 0):.2f}%"
            ])
        
        # Footer
        writer.writerow([])
        writer.writerow([f"Total entries: {len(results)}"])
        
        return output.getvalue().encode('utf-8-sig')
    
    def export_ngram_csv(self, results: List[Dict[str, Any]]) -> bytes:
        """
        Export n-gram data as CSV with watermark.
        
        Args:
            results: List of n-gram data
                     Each dict should have: ngram, frequency, n_value
        
        Returns:
            CSV file content as bytes
        """
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Watermark header
        citation = self.generate_citation()
        for line in citation.split('\n'):
            writer.writerow([f"# {line}"])
        writer.writerow([])
        
        # Column headers
        writer.writerow(['N-gram', 'Frequency', 'N-value'])
        
        # Data rows
        for result in results:
            writer.writerow([
                result.get('ngram', ''),
                result.get('frequency', 0),
                result.get('n_value', 2)
            ])
        
        # Footer
        writer.writerow([])
        writer.writerow([f"Total n-grams: {len(results)}"])
        
        return output.getvalue().encode('utf-8-sig')
    
    def export_ngram_json(self, results: List[Dict[str, Any]]) -> bytes:
        """
        Export n-gram data as JSON with metadata.
        
        Args:
            results: List of n-gram data
        
        Returns:
            JSON file content as bytes
        """
        # Get document title from metadata or filename
        if self.document:
            doc_title = self.document.filename
            if hasattr(self.document, 'metadata') and isinstance(self.document.metadata, dict):
                doc_title = self.document.metadata.get('title', self.document.filename)
        else:
            doc_title = 'All Corpus'
        
        data = {
            'metadata': {
                'platform': 'CorpusLIO National Corpus',
                'exported_by': self.user.username,
                'export_date': self.timestamp.isoformat(),
                'document': doc_title,
                'query': self.query_text,
                'total_ngrams': len(results),
                'citation': self.generate_citation()
            },
            'ngrams': results
        }
        
        return json.dumps(data, ensure_ascii=False, indent=2).encode('utf-8')
    
    def export_concordance_json(self, results: List[Dict[str, Any]]) -> bytes:
        """
        Export concordance results as JSON with metadata.
        
        Args:
            results: List of concordance results
        
        Returns:
            JSON file content as bytes
        """
        # Get document title from metadata or filename
        if self.document:
            doc_title = self.document.filename
            if hasattr(self.document, 'metadata') and isinstance(self.document.metadata, dict):
                doc_title = self.document.metadata.get('title', self.document.filename)
        else:
            doc_title = 'All Corpus'
        
        data = {
            'metadata': {
                'platform': 'CorpusLIO National Corpus',
                'exported_by': self.user.username,
                'export_date': self.timestamp.isoformat(),
                'document': doc_title,
                'query': self.query_text,
                'total_results': len(results),
                'citation': self.generate_citation()
            },
            'results': results
        }
        
        return json.dumps(data, ensure_ascii=False, indent=2).encode('utf-8')
    
    def export_frequency_json(self, results: List[Dict[str, Any]]) -> bytes:
        """
        Export frequency table as JSON with metadata.
        
        Args:
            results: List of frequency data
        
        Returns:
            JSON file content as bytes
        """
        # Get document title from metadata or filename
        if self.document:
            doc_title = self.document.filename
            if hasattr(self.document, 'metadata') and isinstance(self.document.metadata, dict):
                doc_title = self.document.metadata.get('title', self.document.filename)
        else:
            doc_title = 'All Corpus'
        
        data = {
            'metadata': {
                'platform': 'CorpusLIO National Corpus',
                'exported_by': self.user.username,
                'export_date': self.timestamp.isoformat(),
                'document': doc_title,
                'total_entries': len(results),
                'citation': self.generate_citation()
            },
            'frequency_table': results
        }
        
        return json.dumps(data, ensure_ascii=False, indent=2).encode('utf-8')
    
    def export_concordance_excel(self, results: List[Dict[str, Any]]) -> bytes:
        """
        Export concordance results as Excel with watermark header/footer.
        
        Args:
            results: List of concordance results
        
        Returns:
            Excel file content as bytes
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Concordance"
        
        # Style definitions
        header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        watermark_fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
        watermark_font = Font(italic=True, size=9)
        
        # Watermark header (merged cells)
        citation_lines = self.generate_citation().split('\n')
        for idx, line in enumerate(citation_lines, 1):
            ws.merge_cells(f'A{idx}:E{idx}')
            cell = ws[f'A{idx}']
            cell.value = line
            cell.fill = watermark_fill
            cell.font = watermark_font
            cell.alignment = Alignment(horizontal='left', vertical='top')
        
        # Empty row
        watermark_rows = len(citation_lines) + 1
        ws.append([])
        
        # Column headers
        header_row = watermark_rows + 1
        headers = ['Left Context', 'Keyword', 'Right Context', 'Document', 'Position']
        ws.append(headers)
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        # Data rows
        for result in results:
            ws.append([
                result.get('left_context', ''),
                result.get('keyword', ''),
                result.get('right_context', ''),
                result.get('document', ''),
                result.get('position', '')
            ])
        
        # Footer
        ws.append([])
        ws.append([f"Total results: {len(results)}"])
        
        # Column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 10
        
        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()
    
    def export_frequency_excel(self, results: List[Dict[str, Any]]) -> bytes:
        """
        Export frequency table as Excel with watermark.
        
        Args:
            results: List of frequency data
        
        Returns:
            Excel file content as bytes
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Frequency"
        
        # Watermark header
        citation_lines = self.generate_citation().split('\n')
        watermark_fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
        watermark_font = Font(italic=True, size=9)
        
        for idx, line in enumerate(citation_lines, 1):
            ws.merge_cells(f'A{idx}:E{idx}')
            cell = ws[f'A{idx}']
            cell.value = line
            cell.fill = watermark_fill
            cell.font = watermark_font
        
        watermark_rows = len(citation_lines) + 1
        ws.append([])
        
        # Headers
        header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        headers = ['Word', 'Lemma', 'POS', 'Frequency', 'Percentage']
        ws.append(headers)
        
        for col_idx, _ in enumerate(headers, 1):
            cell = ws.cell(row=watermark_rows + 1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
        
        # Data
        for result in results:
            ws.append([
                result.get('word', ''),
                result.get('lemma', ''),
                result.get('pos', ''),
                result.get('frequency', 0),
                f"{result.get('percentage', 0):.2f}%"
            ])
        
        # Column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 12
        
        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()
    
    def export_conllu(self, sentence_indices: List[int] = None) -> bytes:
        """
        Export document as CoNLL-U format with watermark.
        
        Args:
            sentence_indices: Optional list of sentence indices to export.
                            If None, export all sentences.
        
        Returns:
            CoNLL-U formatted text as bytes
            
        Raises:
            ValueError: If document has no dependency annotations
        """
        if not self.document.analysis.has_dependencies:
            raise ValueError("Document has no dependency annotations")
        
        from corpuslio.parsers.conllu_parser import CoNLLUParser
        
        tokens = self.document.analysis.conllu_data or []
        
        if not tokens:
            raise ValueError("No CoNLL-U data available")
        
        # Filter by sentence indices if provided
        if sentence_indices:
            tokens = [t for t in tokens if t.get('sentence_id') in sentence_indices]
        
        # Add watermark as comment lines at the beginning
        citation = self.generate_citation()
        watermark_lines = [
            "# ========================================",
            "# CoNLL-U Export from CorpusLIO Platform",
            "# ========================================",
        ]
        
        for line in citation.split('\n'):
            if line.strip():
                watermark_lines.append(f"# {line}")
        
        watermark_lines.append("# ========================================")
        watermark_lines.append("")
        
        watermark_text = '\n'.join(watermark_lines)
        
        # Serialize tokens to CoNLL-U format
        conllu_text = CoNLLUParser.serialize(tokens, include_metadata=True)
        
        # Combine watermark and content
        full_output = watermark_text + conllu_text
        
        return full_output.encode('utf-8')
